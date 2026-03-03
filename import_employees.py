import sqlite3
from openpyxl import load_workbook

# ✅ correct file path
file_path = r"C:\Users\Administrator\conference\EMP Details Dept-26.xlsx"

# load excel
wb = load_workbook(file_path)
sheet = wb.active

conn = sqlite3.connect("booking.db")
cursor = conn.cursor()

# ✅ ensure users table has correct columns
cursor.execute("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id TEXT UNIQUE,
    username TEXT,
    department TEXT,
    password TEXT,
    role TEXT
)
""")

for row in sheet.iter_rows(min_row=2, values_only=True):
    username = row[1]
    department = row[2]
    emp_id = str(row[3])
    password = str(row[4])

    # auto role
    role = "admin" if department and department.lower() == "admin" else "user"

    cursor.execute("""
        INSERT OR IGNORE INTO users
        (employee_id, username, department, password, role)
        VALUES (?, ?, ?, ?, ?)
    """, (emp_id, username, department, password, role))

conn.commit()
conn.close()

print("✅ Employees imported successfully!")